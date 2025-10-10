import json
from pathlib import Path
from openpyxl import load_workbook, worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill

from . import constants

CENTER_ALIGNED = Alignment(horizontal='center', vertical='center')


def is_person_id_valid(person_id):
    """Вспомогательная функция для проверки валидности ID."""
    return person_id not in {
        constants.PERSON_ID_STATUS_NOT_FOUND,
        constants.PERSON_ID_STATUS_MULTIPLE_FOUND,
        constants.PERSON_ID_STATUS_API_ERROR
    }


def save_json(data, filename):
    with open(filename, 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=2)


# def analyze_person_ids(data, results_path: Path):
#     not_found_file = results_path / "not_found_records.json"
#     doubles_file = results_path / "doubles_records.json"
#
#     not_found = [r for r in data if r['person']['id'] == constants.PERSON_ID_STATUS_NOT_FOUND]
#     if not_found:
#         save_json(not_found, not_found_file)
#
#     doubles = [r for r in data if r['person']['id'] == constants.PERSON_ID_STATUS_MULTIPLE_FOUND]
#     if doubles:
#         save_json(doubles, doubles_file)
#
#     return data


def _align_column_center(sheet: worksheet, columns: list):
    for column_to_align in columns:
        for cell in sheet[column_to_align]:
            cell.alignment = CENTER_ALIGNED


def _align_row_center(sheet: worksheet, rows: list):
    for row_to_align in rows:
        for cell in sheet[row_to_align]:
            cell.alignment = CENTER_ALIGNED


def _auto_cells_width(sheet):
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length + 4


# def doubles_and_not_found(book_path: Path, results_path: Path):
#     process_map = {
#         "Не найдено в ЕВМИАС": results_path / "not_found_records.json",
#         "Двойники": results_path / "doubles_records.json"
#     }
#
#     book = load_workbook(book_path)
#
#     for sheet_name, file_name in process_map.items():
#         if file_name.is_file():
#             sheet = book.create_sheet(sheet_name)
#             sheet.append(["Дата взятия", "ИНЗ", "ФИО", "Дата рождения"])
#             with open(file_name, "r", encoding="utf-8") as f:
#                 data = json.load(f)
#
#             unique_records = set()
#             for each in data:
#                 person = each["person"]
#                 row_tuple = (
#                     each["visit_date"],
#                     each["inz"],
#                     f"{person['last_name']} {person['first_name']} {person['middle_name']}".strip(),
#                     person['birth_day']
#                 )
#                 unique_records.add(row_tuple)
#
#             for record in sorted(list(unique_records)):
#                 sheet.append(record)
#
#             _auto_cells_width(sheet)
#
#             rows_to_align = [1]
#             _align_row_center(sheet, rows_to_align)
#
#             columns_to_align = ["A", "B", "D"]
#             _align_column_center(sheet, columns_to_align)
#
#     book.save(book_path)
#     book.close()


def make_report(data: list[dict], filename: str | Path):
    book = load_workbook(filename)
    if "Для работы" in book.sheetnames:
        book.remove(book["Для работы"])
    sheet = book.create_sheet("Для работы")

    headers = [
        "Дата взятия", "ИНЗ", "ФИО", "Дата рождения", "Код теста",
        "Название теста", "Кол-во", "Цена", "Оплата", "Комментарий"
    ]
    sheet.append(headers)

    invalid_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

    for each in data:
        name = f"{each.get('last_name', '')} {each.get('first_name', '')} {each.get('middle_name', '')}".strip()

        row_to_append = [
            each.get("visit_date", ""),
            each.get("inz", ""),
            name,
            each.get("birth_day", ""),
            each.get("test_code", ""),
            each.get("test_name", ""),
            each.get("test_quantity", 0),
            each.get("test_price", 0.0),
            each.get("test_pay_type", ""),
            each.get("comment", ""),
        ]
        sheet.append(row_to_append)

        # Красим строку, если есть комментарий
        if each.get("comment"):
            current_row = sheet.max_row
            for cell in sheet[current_row]:
                cell.fill = invalid_fill

    # автоширина всех колонок
    _auto_cells_width(sheet)

    # выравнивание строк по центру
    rows_to_align = [1]
    _align_row_center(sheet, rows_to_align)

    # выравнивание колонок по центру
    columns_to_align = ["A", "B", "D"]
    _align_column_center(sheet, columns_to_align)

    # задаем формат ячейки 0,00
    price_column = 'H'
    for cell in sheet[price_column][1:]:
        cell.number_format = constants.PRICE_FORMAT

    book.save(filename)
    book.close()
