from pathlib import Path
from openpyxl import load_workbook
from app.core.logger_setup import logger
from .pay_type_mapper import PAY_TYPE_IDS
from .request import (
    fetch_person_id,
    fetch_test_data_from_evmias,
    fetch_person_tests_history,
    fetch_test_report,
    fetch_medical_history,
    fetch_pay_type_id
)
from .sanitizer import (
    sanitize_test_data_from_evmias,
    sanitize_test_info,
    sanitize_medical_history
)
from app.service.gateway import GatewayService


def get_raw_data(book_path: Path, start_row: int, max_col: int, min_col: int = 2):
    logger.info(f"Обработка файла {book_path}")
    if not book_path.exists():
        logger.error(f"Ошибка: Файл не найден по пути {book_path}")
        return []

    book = load_workbook(book_path)
    sheet = book.active
    visit_date, patient_birthday = None, None
    processed_data, seen_rows = [], set()

    for row in sheet.iter_rows(min_row=start_row, min_col=min_col, max_col=max_col, values_only=True):
        row = [item for item in row if item is not None]
        if not row: break
        if len(row) == 1:
            visit_date = row[0].strftime('%d.%m.%Y') if hasattr(row[0], 'strftime') else str(row[0])
        elif len(row) == 2:
            patient_birthday = row[0].strftime('%d.%m.%Y') if hasattr(row[0], 'strftime') else str(row[0])
        elif len(row) > 2:
            if visit_date and patient_birthday:
                combined_row = [visit_date, patient_birthday] + row
                row_tuple = tuple(combined_row)
                if row_tuple not in seen_rows:
                    processed_data.append(combined_row)
                    seen_rows.add(row_tuple)
    book.close()
    return processed_data


async def get_ids(
        service: GatewayService, data: list,
        task_id: str, manager,
        start_progress: int = 35, end_progress: int = 45
) -> list:
    result = []
    total = len(data)

    if total == 0:
        return []

    progress_span = end_progress - start_progress

    for i, row in enumerate(data):
        record = await fetch_person_id(service, row)
        result.append(record)

        current_progress = start_progress + int(((i + 1) / total) * progress_span)
        progress_message = {
            "progress": current_progress,
            "detail": f"Обработано {i + 1} из {total}"
        }
        await manager.send_progress(task_id, progress_message)

    return result


async def get_test_data_from_evmias(
        service: GatewayService, data: list,
        task_id: str, manager,
        start_progress: int = 75, end_progress: int = 85
) -> list:
    """
    Обогащает записи данными об услугах из ЕВМИАС.
    """
    total = len(data)
    if total == 0:
        return data

    progress_span = end_progress - start_progress

    for i, row in enumerate(data):
        test_code = row['test_src']['code']

        test_data_list = await fetch_test_data_from_evmias(service, test_code)

        found_test = None
        if test_data_list:
            # Ищем точное совпадение по коду в ответе
            for test_item in test_data_list:
                if test_item.get("UslugaComplex_Code") == test_code:
                    found_test = sanitize_test_data_from_evmias(test_item)
                    break

        row['test_evmias'] = found_test

        # Отправляем прогресс
        current_progress = start_progress + int(((i + 1) / total) * progress_span)
        progress_message = {
            "progress": current_progress,
            "detail": f"Поиск информации об услугах: {i + 1} из {total}"
        }
        await manager.send_progress(task_id, progress_message)

    return data


async def get_person_tests_history(
        service: GatewayService, data: list,
        task_id: str, manager,
        start_progress: int, end_progress: int
) -> list:
    """
    Обогащает записи историей лабораторных исследований ('lab') пациента.
    """
    total = len(data)
    if total == 0:
        return data

    progress_span = end_progress - start_progress

    for i, row in enumerate(data):
        person_id = row["person"]["id"]

        test_history_raw = await fetch_person_tests_history(service, person_id)

        filtered_history = [
            item for item in test_history_raw
            if 'lab' in item.get("UslugaComplex_AttributeList", "")
        ]

        row['tests_history'] = filtered_history

        current_progress = start_progress + int(((i + 1) / total) * progress_span)
        progress_message = {
            "progress": current_progress,
            "detail": f"Получение истории анализов: {i + 1} из {total}"
        }
        await manager.send_progress(task_id, progress_message)

    return data


async def get_pay_type(
        service: GatewayService, data: list,
        task_id: str, manager,
        start_progress: int, end_progress: int
) -> list:
    """
    Обогащает записи данными о типе оплаты, полученными из отчета по анализу.
    """
    total = len(data)
    if total == 0:
        return data

    progress_span = end_progress - start_progress

    for i, row in enumerate(data):
        # Бизнес-логика: выполняем только если есть данные о тесте и история
        if row.get("test_evmias") is not None and row.get("tests_history"):
            tests_history = row["tests_history"]

            # Извлекаем ID события. `tests_history` может быть списком или одним элементом.
            # На будущих этапах он будет отфильтрован до одного, но сейчас лучше сделать проверку.
            test_id = None
            if isinstance(tests_history, list) and tests_history:
                test_id = tests_history[0].get("event_id")
            elif isinstance(tests_history, dict):
                test_id = tests_history.get("event_id")

            if test_id:
                test_report_raw = await fetch_test_report(service, test_id)
                # Безопасно извлекаем вложенные данные
                if test_report_raw:
                    report_data = test_report_raw.get("map", {}).get("EvnUslugaPar", {}).get("item", [{}])[0].get(
                        "data")
                    if report_data:
                        row["test_report"] = sanitize_test_info(report_data)
                    else:
                        row["test_report"] = None
                else:
                    row["test_report"] = None
            else:
                row["test_report"] = None
        else:
            row["test_report"] = None

        # Отправляем прогресс
        current_progress = start_progress + int(((i + 1) / total) * progress_span)
        progress_message = {
            "progress": current_progress,
            "detail": f"Поиск в истории анализов: {i + 1} из {total}"
        }
        await manager.send_progress(task_id, progress_message)

    return data


async def get_medical_history(
        service: GatewayService, data: list,
        task_id: str, manager,
        start_progress: int, end_progress: int
) -> list:
    """
    Запасной механизм: для записей, где не удалось найти тип оплаты,
    пытается найти его через общую медицинскую историю.
    """
    total = len(data)
    if total == 0:
        return data

    progress_span = end_progress - start_progress

    for i, row in enumerate(data):
        # Логика: выполняем только если `test_report` еще не заполнен
        if not row.get("test_report"):
            person_id = row["person"]["id"]
            visit_date = row["visit_date"]

            medical_history_raw = await fetch_medical_history(service, person_id)

            # Санитизация и фильтрация истории по дате визита (+- 14 дней)
            sanitized_history = sanitize_medical_history(medical_history_raw, visit_date)
            row["medical_history"] = sanitized_history

            # Если в отфильтрованной истории что-то нашлось
            if sanitized_history:
                # Берем первое подходящее событие
                first_event = sanitized_history[0]
                event_id = first_event.get("children_evn_id")
                med_staff_fact_id = first_event.get("med_staff_fact_id")

                if event_id:
                    pay_type_id = await fetch_pay_type_id(service, event_id)
                    if pay_type_id:
                        row["test_report"] = {
                            "pay_type_id": pay_type_id,
                            "pay_type": PAY_TYPE_IDS.get(pay_type_id, "Неизвестно"),
                            "med_staff_fact_id": med_staff_fact_id
                        }

        # Отправляем прогресс в любом случае
        current_progress = start_progress + int(((i + 1) / total) * progress_span)
        progress_message = {
            "progress": current_progress,
            "detail": f"Поиск в мед. истории: {i + 1} из {total}"
        }
        await manager.send_progress(task_id, progress_message)

    return data
